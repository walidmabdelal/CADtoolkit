# -*- coding: utf-8 -*-
"""
Created on Tue Oct  2 08:54:44 2018

@author: Walid Abdelal
"""

import win32com.client,os, os.path, tkinter.filedialog
from tkinter import messagebox
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import string


# Get a COM object for autocad
acad = win32com.client.Dispatch("AutoCAD.Application")
# Initiate local variables
today = datetime.today().strftime('%Y-%m-%d, %H:%M:%S')
xrefs_paths = []
xrefs_stats = []
report_fields = ['No', 'F_path', 'Title', 'Modified', 'Xrefs_paths', 
                 'Xrefs_names', 'Xref modified', 'file name modified', 'Date']
modified = 0
letters = string.ascii_uppercase
specialChars = "~#%&*}{\:<>?/|"
validFilenameChars = "-_.()@ %s%s" % (string.ascii_letters, string.digits)

# Functions definitions
def repath(filename):
    print ('\n\n Repathing %s...' %filename)
    doc = acad.Documents.Open(filename)
    
    blocks = doc.Database.Blocks # Internally xrefs are just blocks!
    xrefs = [item for item in blocks if item.IsXRef]
    
    if xrefs:
        for xref in xrefs:
            old_path = xref.Path
            print('Repathing...' + xref.name)
            new_path = os.path.join('.\\Test\\', os.path.basename(old_path))
            xref.Path = new_path
            time.sleep(1)
            print ('Old path name was %s, new path name is %s.\n' 
                   %(old_path, new_path))
    try:
        doc.Close(True) # Close and save
    except: # Something went wrong,
        doc.Close(False) # close then report it
        raise
        
def reXref(filename):
    print ('\n\n Checking %s...' %filename)
    doc = acad.Documents.Open(filename)
    
    blocks = doc.Database.Blocks # Internally xrefs are just blocks!
    xrefs = [item for item in blocks if item.IsXRef]
    modified = 0 # Default is not touched
    
    if xrefs:
        for xref in xrefs:
            xref_path = xref.Path
            xrefs_paths.append(xref_path) #populate xrefs paths array
            xref_name = os.path.basename(xref.Path)
            # Cleaning xref path from ambersands
            print('Checking...' + xref_name)
            
            if any(c in xref_path for c in specialChars):
                if xref_path.find('&') != -1:
                    xref.Path = xref_path.replace('&', 'and')
                xref.Path = ''.join(c for c in xref_path 
                                    if c in validFilenameChars)
                xrefs_paths[len(xrefs_paths) - 1] = xref.Path
                xrefs_stats.append(1) # Xref status flag
                modified = 1 # Modified state tracker
                print ('\n\'&\' was removed from %s.\n' %xref.Path)
                
            else:
                xrefs_stats.append(0) # Xref status flag
                print ('\n no operations requried')
                
            time.sleep(1)
    try:
        doc.Close(True) # Close and save
    except: # Something went wrong,
        doc.Close(False) # close then report it
        raise
    return modified

# Cleaning the physical xref file name from the ambersand    
def cleanse_xref_name(xref_path):
    if any(c in xref_path for c in specialChars):
        xref_modified = 1
        clean_xref_path = xref_path.replace('&', 'and') 
        clean_xref_path = ''.join(c for c in clean_xref_path 
                                    if c in validFilenameChars)
        os.rename(xref_path, clean_xref_path)
        print('\n Renaming file...\n' + os.path.basename(xref_path) +
              'was renamed to' + os.path.basename(clean_xref_path))
    else:
        xref_modified = 0
    return xref_modified


def gen_report(cad_f, report, flag, f_flag, isdgn):
    
    if os.path.isfile(dir + '\\Xrefs.xlsx'):
        report = load_workbook(dir + '\\Xrefs.xlsx', read_only=False)
    #ws = report.create_sheet(today)
    ws = report.active
    if ws['A1'] != 'No':
        for cll in range(1, 10):         
            ws.cell(1, cll).value = report_fields[cll - 1]
            if ws.cell(1, cll).coordinate == ['K1']:
                break
    # Write metadata
    last_row = ws.max_row
    idx = str(last_row + 1)
    ws['A' + idx].value = last_row
    ws['B' + idx].value = cad_f
    ws['C' + idx].value = os.path.basename(cad_f)
    if isdgn:
        ws['C' + idx].fill = PatternFill(start_color="FFFF33", 
                                          end_color="FFFF33", 
                                          fill_type="solid")
    if flag == 1:
        ws['D' + idx].value = 'Y'
        ws['D' + idx].fill = PatternFill(start_color="FF0000", 
                                          end_color="FF0000", 
                                          fill_type="solid")
    else:
        ws['D' + idx].value = 'N'
        ws['B' + idx].fill = PatternFill(start_color="008000", 
                                          end_color="008000", 
                                         fill_type="solid")
    if f_flag == 1:
        ws['H' + idx].value = 'Y'
        ws['H' + idx].fill = PatternFill(start_color="FF0000", 
                                          end_color="FF0000", 
                                          fill_type="solid")
    else:
        ws['H' + idx].value = 'N'
        ws['H' + idx].fill = PatternFill(start_color="008000", 
                                          end_color="008000", 
                                         fill_type="solid")  
    ws['I' + idx].value = today
    # Write associated X-refs
    if not isdgn:
        for i in range(len(xrefs_paths)):
            ws['E' + str(int(idx) + i)].value = xrefs_paths[i-1]
            ws['F' + str(int(idx) + i)].value = os.path.basename(xrefs_paths[i-1])
            if xrefs_stats[i] == 1:
               ws['E' + str(int(idx) + i)].fill = PatternFill(start_color="FF0000", 
                                              end_color="FF0000", 
                                              fill_type="solid")
               ws['G' + str(int(idx) + i)].value = 'Y'
            else:
               ws['G' + str(int(idx) + i)].value = 'N' 
    report.save(dir + '\\Xrefs.xlsx')
        
# Main program
if __name__ == '__main__':
    if acad.Visible:
        acad.Visible = False
    time.sleep(1)
    
    # Get working directory
    dir = tkinter.filedialog.askdirectory()
    answer = messagebox.askokcancel('RePath','Re path all dwg files in ' + dir + '?')
    
    # Prepare report file
    if os.path.isfile(dir + '\\Xrefs.xlsx'):
        print('\nExisting report found')
        report = load_workbook(dir + '\\Xrefs.xlsx', read_only=False)
        
    else:
        print('\nNew report file generated')
        report = Workbook(write_only=False)
        
    if answer:
        for dirpath, subdirs, files in os.walk(dir):
            for name in files:
                ext = name.split('.')[-1] or ''
                # We want dwg files which are not in the x-ref directory
                if ext.lower() == 'dwg':
                    drawing_path = os.path.join(dirpath, name)
                    del xrefs_paths[:]
                    del xrefs_stats[:]
                    try:
                        flag = reXref(drawing_path)
                        f_flag = cleanse_xref_name(drawing_path)
                        gen_report(drawing_path, report, flag, f_flag, isdgn=0)
                        print('\nSuccessful')
                    except:
                        print ('\nUnable to repath drawing %s!' %drawing_path)
                elif ext.lower() == 'dgn':
                    drawing_path = os.path.join(dirpath, name)
                    flag = 0
                    f_flag = 0
                    try:
                        gen_report(drawing_path, report, flag, f_flag, isdgn=0)
                        print('\nDGN file skipped')
                    except:
                        print ('\nUnable to generate report')
                else: 
                    print('\nNot a CAD file.\n')
    acad.Visible = True
    
else: print('Please run me from the source code...')